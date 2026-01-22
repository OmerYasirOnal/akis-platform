#!/usr/bin/env python3
"""
AKIS Platform - Planning Spreadsheet Generator

Creates AKIS_Planning_WBS.xlsx based on repository audit outcomes.
Preserves: filters, frozen panes, column names, table styling, conditional formatting.

Usage: python3 generate_planning_spreadsheet.py
Output: docs/planning/AKIS_Planning_WBS.xlsx
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Constants
OUTPUT_PATH = "docs/planning/AKIS_Planning_WBS.xlsx"
AUDIT_DATE = datetime(2026, 1, 13)

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
COMPLETED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PLANNED_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook():
    """Create workbook with all required tabs."""
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Create tabs in order
    create_overview_tab(wb)
    create_kisiler_tab(wb)
    create_fazlar_tab(wb)
    create_sprintler_tab(wb)
    create_gorevler_tab(wb)
    create_gantt_fazlar_tab(wb)
    create_changelog_tab(wb)
    
    # Remove the default empty sheet
    wb.remove(default_sheet)
    
    return wb

def apply_header_style(row, num_cols):
    """Apply header styling to a row."""
    for cell in row[:num_cols]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def auto_width(ws, num_cols):
    """Auto-adjust column widths."""
    for i in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

def create_overview_tab(wb):
    """00_Overview tab - Project summary."""
    ws = wb.create_sheet("00_Overview")
    
    headers = ["Proje Adı", "Versiyon", "Güncelleme Tarihi", "Durum", "Açıklama"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    ws.append([
        "AKIS Platform",
        "1.0.0",
        AUDIT_DATE.strftime("%Y-%m-%d"),
        "STAGING-READY",
        "AI Agent Workflow Engine - Bitirme Projesi"
    ])
    
    # Summary section
    ws.append([])
    ws.append(["ÖZET İSTATİSTİKLER"])
    ws["A4"].font = Font(bold=True, size=12)
    
    ws.append(["Metrik", "Değer"])
    ws.append(["Toplam Görev", "=COUNTA('04_Gorevler'!A:A)-1"])
    ws.append(["Tamamlanan", "=COUNTIF('04_Gorevler'!F:F,\"TAMAMLANDI\")"])
    ws.append(["Devam Eden", "=COUNTIF('04_Gorevler'!F:F,\"DEVAM\")"])
    ws.append(["Planlı", "=COUNTIF('04_Gorevler'!F:F,\"PLANLI\")"])
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

def create_kisiler_tab(wb):
    """01_Kisiler tab - Team members."""
    ws = wb.create_sheet("01_Kisiler")
    
    headers = ["Kişi_ID", "Ad Soyad", "Rol", "E-posta", "Notlar"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    team = [
        ["K-001", "Ömer Yasir Önal", "Proje Sahibi / Geliştirici", "omer@example.com", "Bitirme projesi öğrencisi"],
        ["K-002", "Danışman", "Akademik Danışman", "-", "Tez danışmanı"],
    ]
    for row in team:
        ws.append(row)
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

def create_fazlar_tab(wb):
    """02_Fazlar tab - Project phases."""
    ws = wb.create_sheet("02_Fazlar")
    
    headers = ["Faz_ID", "Faz Adı", "Başlangıç", "Bitiş", "Durum", "Açıklama"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    phases = [
        ["F-1", "Temel Altyapı", "2025-10-01", "2025-11-30", "TAMAMLANDI", "Backend, Frontend, DB setup"],
        ["F-2", "Auth & OAuth", "2025-12-01", "2025-12-20", "TAMAMLANDI", "JWT, Multi-step auth, GitHub/Google OAuth"],
        ["F-3", "Agent Orchestration", "2025-12-21", "2026-01-05", "TAMAMLANDI", "Scribe agent, MCP adapters"],
        ["F-4", "UI Modernization", "2026-01-06", "2026-01-12", "TAMAMLANDI", "Cursor-inspired dashboard, Integrations Hub"],
        ["F-5", "Staging Deployment", "2026-01-13", "2026-01-20", "DEVAM", "OCI Free Tier setup, CI/CD"],
        ["F-6", "Production Launch", "2026-01-21", "2026-01-31", "PLANLI", "Production deployment, monitoring"],
    ]
    for row in phases:
        ws.append(row)
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

def create_sprintler_tab(wb):
    """03_Sprintler tab - Sprint planning."""
    ws = wb.create_sheet("03_Sprintler")
    
    headers = ["Sprint_ID", "Sprint Adı", "Faz_ID", "Başlangıç", "Bitiş", "Durum", "Hedef"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    sprints = [
        ["S1.1", "Backend Core", "F-1", "2025-10-01", "2025-10-15", "TAMAMLANDI", "Fastify server, PostgreSQL"],
        ["S1.2", "Frontend Core", "F-1", "2025-10-16", "2025-10-31", "TAMAMLANDI", "React SPA, Vite setup"],
        ["S1.3", "CI/CD Setup", "F-1", "2025-11-01", "2025-11-30", "TAMAMLANDI", "GitHub Actions, PR gates"],
        ["S2.1", "Auth Core", "F-2", "2025-12-01", "2025-12-10", "TAMAMLANDI", "JWT, multi-step auth"],
        ["S2.2", "OAuth Integration", "F-2", "2025-12-11", "2025-12-20", "TAMAMLANDI", "GitHub, Google, Atlassian OAuth"],
        ["S3.1", "Scribe Agent", "F-3", "2025-12-21", "2026-01-05", "TAMAMLANDI", "Scribe core, dry-run, write mode"],
        ["S4.1", "Dashboard UI", "F-4", "2026-01-06", "2026-01-12", "TAMAMLANDI", "Cursor-style UI, Integrations Hub"],
        ["S5.1", "OCI Staging", "F-5", "2026-01-13", "2026-01-20", "DEVAM", "VM provisioning, deploy workflow"],
        ["S6.1", "Production", "F-6", "2026-01-21", "2026-01-31", "PLANLI", "Prod deploy, monitoring"],
    ]
    for row in sprints:
        ws.append(row)
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

def create_gorevler_tab(wb):
    """04_Gorevler tab - Task list with WBS IDs."""
    ws = wb.create_sheet("04_Gorevler")
    
    headers = ["Görev_ID", "Görev Adı", "Sprint_ID", "Workstream", "Sorumlu", "Durum", 
               "Planlanan Başlangıç", "Planlanan Bitiş", "Gerçek Başlangıç", "Gerçek Bitiş", 
               "Kanıt", "Notlar"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    # Task data based on audit
    tasks = [
        # Phase 1 - Infrastructure
        ["T-INF-01", "Fastify Server Setup", "S1.1", "Backend", "K-001", "TAMAMLANDI", 
         "2025-10-01", "2025-10-07", "2025-10-01", "2025-10-05", "server.app.ts", ""],
        ["T-INF-02", "PostgreSQL + Drizzle ORM", "S1.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-10-08", "2025-10-15", "2025-10-06", "2025-10-12", "drizzle.config.ts", "22 migrations"],
        ["T-INF-03", "React SPA + Vite", "S1.2", "Frontend", "K-001", "TAMAMLANDI",
         "2025-10-16", "2025-10-31", "2025-10-16", "2025-10-28", "vite.config.ts", ""],
        ["T-INF-04", "GitHub Actions CI", "S1.3", "DevOps", "K-001", "TAMAMLANDI",
         "2025-11-01", "2025-11-15", "2025-11-01", "2025-11-10", "ci.yml", ""],
        ["T-INF-05", "PR Gate Workflow", "S1.3", "DevOps", "K-001", "TAMAMLANDI",
         "2025-11-16", "2025-11-30", "2025-11-11", "2025-11-20", "pr-gate.yml", ""],
        
        # Phase 2 - Auth
        ["T-AUTH-01", "JWT Authentication", "S2.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-01", "2025-12-05", "2025-12-01", "2025-12-04", "jwt.ts", ""],
        ["T-AUTH-02", "Multi-step Auth", "S2.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-05", "2025-12-10", "2025-12-04", "2025-12-08", "auth.multi-step.ts", ""],
        ["T-AUTH-03", "GitHub OAuth", "S2.2", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-11", "2025-12-15", "2025-12-09", "2025-12-13", "auth.oauth.ts", "PR #90"],
        ["T-AUTH-04", "Google OAuth", "S2.2", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-15", "2025-12-17", "2025-12-14", "2025-12-16", "auth.oauth.ts", "PR #90"],
        ["T-AUTH-05", "Atlassian OAuth", "S2.2", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-18", "2025-12-20", "2026-01-09", "2026-01-10", "PR #172", "3LO OAuth"],
        
        # Phase 3 - Agents
        ["T-AGT-01", "Agent Orchestrator", "S3.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-21", "2025-12-25", "2025-12-21", "2025-12-24", "AgentOrchestrator.ts", ""],
        ["T-AGT-02", "Scribe Agent Core", "S3.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-25", "2025-12-30", "2025-12-25", "2025-12-29", "scribe/", ""],
        ["T-AGT-03", "MCP Adapters", "S3.1", "Backend", "K-001", "TAMAMLANDI",
         "2025-12-30", "2026-01-05", "2025-12-30", "2026-01-03", "mcp/adapters/", "GitHub/Jira/Confluence"],
        
        # Phase 4 - UI
        ["T-UI-01", "Cursor-inspired Dashboard", "S4.1", "Frontend", "K-001", "TAMAMLANDI",
         "2026-01-06", "2026-01-08", "2026-01-06", "2026-01-07", "PR #166", "Liquid neon theme"],
        ["T-UI-02", "Agents Hub", "S4.1", "Frontend", "K-001", "TAMAMLANDI",
         "2026-01-08", "2026-01-10", "2026-01-07", "2026-01-09", "PR #167", ""],
        ["T-UI-03", "Integrations Hub", "S4.1", "Frontend", "K-001", "TAMAMLANDI",
         "2026-01-10", "2026-01-12", "2026-01-09", "2026-01-10", "PR #169", ""],
        ["T-UI-04", "Live Progress UI", "S4.1", "Frontend", "K-001", "TAMAMLANDI",
         "2026-01-10", "2026-01-12", "2026-01-10", "2026-01-12", "PR #170", "Real-time trace events"],
        
        # Phase 5 - Staging (Current)
        ["T-OCI-01", "OCI VM Provisioning", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-13", "2026-01-14", "", "", "", "ARM64 Free Tier"],
        ["T-OCI-02", "DNS Configuration", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-14", "2026-01-14", "", "", "", "staging.akisflow.com"],
        ["T-OCI-03", "GitHub Secrets Setup", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-14", "2026-01-15", "", "", "", "STAGING_* secrets"],
        ["T-OCI-04", "First Staging Deploy", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-15", "2026-01-16", "", "", "", "Via oci-staging-deploy.yml"],
        ["T-OCI-05", "Smoke Test Execution", "S5.1", "QA", "K-001", "PLANLI",
         "2026-01-16", "2026-01-17", "", "", "", "QA_EVIDENCE_STAGING_SMOKE_PACK.md"],
        ["T-OCI-06", "Staging OAuth Apps", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-17", "2026-01-18", "", "", "", "GitHub/Google OAuth for staging"],
        ["T-OCI-07", "Email Provider (Resend)", "S5.1", "Backend", "K-001", "PLANLI",
         "2026-01-18", "2026-01-19", "", "", "", "Production email delivery"],
        ["T-OCI-08", "Backup Automation", "S5.1", "DevOps", "K-001", "PLANLI",
         "2026-01-19", "2026-01-20", "", "", "", "OCI Object Storage"],
        
        # Phase 6 - Production
        ["T-PROD-01", "Production OAuth Apps", "S6.1", "DevOps", "K-001", "PLANLI",
         "2026-01-21", "2026-01-22", "", "", "", "Separate from staging"],
        ["T-PROD-02", "Production Deploy", "S6.1", "DevOps", "K-001", "PLANLI",
         "2026-01-22", "2026-01-24", "", "", "", "Via deploy-prod.yml"],
        ["T-PROD-03", "Monitoring Setup", "S6.1", "DevOps", "K-001", "PLANLI",
         "2026-01-24", "2026-01-26", "", "", "", "UptimeRobot, alerts"],
        ["T-PROD-04", "Documentation Final", "S6.1", "Docs", "K-001", "PLANLI",
         "2026-01-26", "2026-01-28", "", "", "", "User guide, API docs"],
        ["T-PROD-05", "Demo Presentation", "S6.1", "QA", "K-001", "PLANLI",
         "2026-01-28", "2026-01-31", "", "", "", "Final project demo"],
    ]
    
    for row in tasks:
        ws.append(row)
    
    # Conditional formatting for status
    ws.conditional_formatting.add('F2:F100',
        FormulaRule(formula=['$F2="TAMAMLANDI"'], fill=COMPLETED_FILL))
    ws.conditional_formatting.add('F2:F100',
        FormulaRule(formula=['$F2="DEVAM"'], fill=IN_PROGRESS_FILL))
    ws.conditional_formatting.add('F2:F100',
        FormulaRule(formula=['$F2="PLANLI"'], fill=PLANNED_FILL))
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "B2"
    
    # Set filter
    ws.auto_filter.ref = f"A1:L{len(tasks) + 1}"

def create_gantt_fazlar_tab(wb):
    """05_Gantt_Fazlar tab - Simplified Gantt view."""
    ws = wb.create_sheet("05_Gantt_Fazlar")
    
    headers = ["Faz", "W1", "W2", "W3", "W4", "W5", "W6", "W7", "W8", "W9", "W10", "W11", "W12"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    # Simple Gantt representation
    gantt = [
        ["F-1 Temel Altyapı", "████", "████", "████", "████", "████", "████", "████", "████", "", "", "", ""],
        ["F-2 Auth & OAuth", "", "", "", "", "", "", "", "", "████", "████", "████", ""],
        ["F-3 Agent Orchestration", "", "", "", "", "", "", "", "", "", "", "████", "████"],
        ["F-4 UI Modernization", "", "", "", "", "", "", "", "", "", "", "", "████"],
        ["F-5 Staging Deploy", "", "", "", "", "", "", "", "", "", "", "", "⏳"],
        ["F-6 Production Launch", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    
    for row in gantt:
        ws.append(row)
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "B2"

def create_changelog_tab(wb):
    """06_ChangeLog tab - Audit changes."""
    ws = wb.create_sheet("06_ChangeLog")
    
    headers = ["Tarih", "Görev_ID", "Önceki Durum", "Yeni Durum", "Sebep", "Kanıt"]
    ws.append(headers)
    apply_header_style(ws[1], len(headers))
    
    changes = [
        # Initial audit - 2026-01-13
        ["2026-01-13", "T-AUTH-01", "DEVAM", "TAMAMLANDI", "JWT auth working", "jwt.ts verified"],
        ["2026-01-13", "T-AUTH-02", "DEVAM", "TAMAMLANDI", "Multi-step auth merged", "PR #90"],
        ["2026-01-13", "T-AUTH-03", "DEVAM", "TAMAMLANDI", "GitHub OAuth working", "auth.oauth.ts"],
        ["2026-01-13", "T-AUTH-04", "DEVAM", "TAMAMLANDI", "Google OAuth working", "auth.oauth.ts"],
        ["2026-01-13", "T-AUTH-05", "PLANLI", "TAMAMLANDI", "Atlassian OAuth merged", "PR #172"],
        ["2026-01-13", "T-AGT-01", "DEVAM", "TAMAMLANDI", "Orchestrator complete", "AgentOrchestrator.ts"],
        ["2026-01-13", "T-AGT-02", "DEVAM", "TAMAMLANDI", "Scribe agent working", "Dry-run verified"],
        ["2026-01-13", "T-AGT-03", "DEVAM", "TAMAMLANDI", "MCP adapters ready", "mcp/adapters/"],
        ["2026-01-13", "T-UI-01", "DEVAM", "TAMAMLANDI", "Dashboard UI shipped", "PR #166"],
        ["2026-01-13", "T-UI-02", "DEVAM", "TAMAMLANDI", "Agents Hub complete", "PR #167"],
        ["2026-01-13", "T-UI-03", "DEVAM", "TAMAMLANDI", "Integrations Hub shipped", "PR #169"],
        ["2026-01-13", "T-UI-04", "DEVAM", "TAMAMLANDI", "Live progress working", "PR #170"],
        ["2026-01-13", "T-OCI-01", "-", "PLANLI", "Added for staging", "STAGING_READINESS_REPO_AUDIT.md"],
        ["2026-01-13", "T-OCI-02", "-", "PLANLI", "Added for staging", "DNS needed for staging"],
        ["2026-01-13", "T-OCI-03", "-", "PLANLI", "Added for staging", "GitHub Secrets needed"],
        ["2026-01-13", "T-OCI-04", "-", "PLANLI", "Added for staging", "First deploy task"],
        ["2026-01-13", "T-OCI-05", "-", "PLANLI", "Added for staging", "QA smoke test"],
        ["2026-01-13", "T-OCI-06", "-", "PLANLI", "Added for staging", "OAuth apps for staging"],
        ["2026-01-13", "T-OCI-07", "-", "PLANLI", "Added for staging", "Email provider setup"],
        ["2026-01-13", "T-OCI-08", "-", "PLANLI", "Added for staging", "Backup automation"],
        # Documentation alignment - 2026-01-15
        ["2026-01-15", "DOCS", "-", "-", "Domain aligned to akisflow.com", "STAGING_READINESS_REPO_AUDIT.md"],
        ["2026-01-15", "DOCS", "-", "-", "Cookie name set to akis_session", "AKISFLOW_DOMAIN_STRATEGY.md"],
        ["2026-01-15", "DOCS", "-", "-", "Added Merge/Rollout order section", "OCI_STAGING_RUNBOOK.md"],
        ["2026-01-15", "DOCS", "-", "-", "QA smoke pack cookie aligned", "QA_EVIDENCE_STAGING_SMOKE_PACK.md"],
        ["2026-01-15", "T-OCI-04", "-", "-", "Workflow made manual-only safe", "oci-staging-deploy.yml"],
    ]
    
    for row in changes:
        ws.append(row)
    
    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"

def main():
    """Main entry point."""
    # Create output directory
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    
    # Generate workbook
    wb = create_workbook()
    
    # Save workbook
    wb.save(OUTPUT_PATH)
    print(f"✅ Created: {OUTPUT_PATH}")
    
    # Print summary
    print(f"\nTabs created:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == "__main__":
    main()
